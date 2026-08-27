import os
import pandas as pd
import numpy as np
import unicodedata
from flask import Blueprint, render_template, request, session, redirect, url_for
from sistemas import login_requerido
from logs import guardar_log_compras
from datetime import datetime
from datetime import timedelta
import uuid
#import traceback
import psycopg2
import redis
import json
from io import StringIO
import re
import unicodedata
from psycopg2.extras import RealDictCursor
from pathlib import Path
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from io import BytesIO


SUCURSAL_MAP = {
    "Total Empresa": "CO01,CO02,CO04,CO05,CO06,CO07,CO08,CO09,CO10,CO11,CO12,CO14,CO15,CO16,CO17,CO18,CO19,CO20,CO21,CO22,CO23,CO24,CO25,CO26,CO27,CO28,CO29,MA02",
    "Total Empresa - Mayorista": "CO05,CO09,CO12,CO15,CO21,CO29,MA02",
    "Jujuy - Mayorista": "CO05,CO12,CO15,MA02",
    "Salta - Mayorista": "CO09,CO29,CO21",
    "Oran - Mayorista": "CO21",
    "Total Empresa Minorista": "CO01,CO02,CO04,CO06,CO07,CO08,CO10,CO11,CO14,CO16,CO17,CO18,CO19,CO20,CO22,CO23,CO26,CO27,CO28",
    "Jujuy - Minoristas": "CO01,CO02,CO04,CO06,CO07,CO08,CO10,CO11,CO14,CO16,CO17,CO19,CO20,CO22,CO28",
    "Salta - Minoristas": "CO18,CO23",
    "Jujuy, Salta - Minoritas": "CO01,CO02,CO04,CO06,CO07,CO08,CO10,CO11,CO14,CO16,CO17,CO18,CO19,CO20,CO22,CO23,CO28",
    "Jujuy, Salta - Minoritas y Mayoristas": "CO01,CO02,CO04,CO05,CO06,CO07,CO08,CO09,CO10,CO11,CO12,CO14,CO15,CO16,CO17,CO18,CO19,CO20,CO21,CO22,CO23,CO28,CO29,MA02",
    "Jujuy - Minorista Y Mayorista": "CO01,CO02,CO04,CO05,CO06,CO07,CO08,CO10,CO11,CO12,CO14,CO15,CO16,CO17,CO19,CO20,CO22,CO28,MA02",
    "Tucuman": "CO24,CO25,CO26,CO27",
    "Tucuman - Minorista": "CO26,CO27",
    "Tucuman - Mayorista": "CO24,CO25",
    "Jujuy Capital": "CO01,CO02,CO05,CO07,CO11,CO16,CO19,CO20,CO22"
}


# ---------------- REDIS ----------------

redis_client = redis.Redis(
    host=os.getenv("REDIS_HOST", "localhost"),
    port=6379,
    db=0,
    decode_responses=True
)

def get_db_connection():
    return psycopg2.connect(
        host=os.getenv("DB_HOST", "postgres"),
        port=os.getenv("DB_PORT", "5432"),
        dbname=os.getenv("DB_NAME", "sip"),
        user=os.getenv("DB_USER", "sip_user"),
        password=os.getenv("DB_PASSWORD", "Alberdi2026_db")
    )

def guardar_temporal(lote_id, df):
    redis_client.setex(
        lote_id,
        3600,  # expira en 1 hora
        df.to_json(orient="records")
    )

def recuperar_temporal(lote_id):
    data = redis_client.get(lote_id)
    if data:
        return pd.read_json(data, orient="records")
    return None



#DB_PATH = os.path.join(os.path.dirname(__file__), "sip.s3db")

compras_bp = Blueprint("compras", __name__, url_prefix="/compras")
RUTA_MATERIAL = "/mnt/excel/ARCHIVOS IMPORTANTES/Base de datos completa.xlsx"
HEADERS = ["CODIGO", "EAN", "DESCRIPCION", "Normal", "Oferta", "cenefa", "desde", "hasta", "sucursales", "CÓD. SUCURSALES"]

ALIAS = {
    "CODIGO": ["CODIGO", "codigo", "id", "cod", "material", "mat", "Cód.", "CODGO"],
    "EAN": ["ean", "EAN","codigo ean"],
    "DESCRIPCION": ["descripcion", "DESCRIPCION", "Descripción", "Descrip", "nombre", "texto breve de material", "Texto breve de material"],
    "Normal": ["normal", "precio normal", "precio unitario", "Precio", "PVN", "pvn", "Nrmal"],
    "Oferta": ["oferta"],
    "cenefa": ["cenefa", "cenefas", "Cenefas","Promo","promo","PROMO"],
    "desde": ["desde", "inicio"],
    "hasta": ["hasta", "fin"],
    "sucursales": ["sucursales", "tiendas", "sap", "SAP", "Sap"],
    "CÓD. SUCURSALES": ["codigos sucursales", "cod sucursales", "sucursales codigos"],
}

SUCURSAL_MAP = {
    "minorista": {
        "Total-empresa-minorista": "CO01,CO02,CO04,CO06,CO07,CO08,CO10,CO11,CO14,CO16,CO17,CO18,CO19,CO20,CO22,CO23,CO26,CO27,CO28",
        "tucuman": "CO26,CO27",
        "jujuy": "CO01,CO02,CO04,CO06,CO07,CO08,CO10,CO11,CO14,CO16,CO17,CO19,CO20,CO22,CO28",
        "salta": "CO18,CO23",
        "tucuman":"CO26,CO27"
    },
    "mayorista": {
        "Total Empresa - Mayorista": "CO05,CO09,CO12,CO15,CO21,CO24,CO25,CO29,MA02",
        "jujuy": "CO05,CO12,CO15,MA02",
        "salta": "CO09,CO29,CO21",
        "oran": "CO21",
        "tucuman":"CO24,CO25"
    }
}

def crear_dataframe_solo_valores(
    archivo,
    hoja_objetivo,
    fila_header,
):
    """
    Lee el Excel usando los valores calculados
    almacenados en el archivo.

    Las fórmulas no se copian al DataFrame:
    se toma únicamente su último resultado calculado.
    """

    archivo.seek(0)
    contenido = archivo.read()

    buffer_valores = BytesIO(contenido)

    workbook = load_workbook(
        buffer_valores,
        data_only=True,
        read_only=True,
    )

    if hoja_objetivo not in workbook.sheetnames:
        workbook.close()
        archivo.seek(0)

        raise ValueError(
            f"No existe la hoja '{hoja_objetivo}'."
        )

    hoja = workbook[hoja_objetivo]

    filas = list(
        hoja.iter_rows(
            values_only=True
        )
    )

    workbook.close()
    archivo.seek(0)

    if not filas:
        raise ValueError(
            "La hoja seleccionada está vacía."
        )

    if fila_header >= len(filas):
        raise ValueError(
            "No se pudo localizar correctamente "
            "la fila de encabezados."
        )

    encabezados_originales = list(
        filas[fila_header]
    )

    # ---------------------------------------------
    # NORMALIZAR ENCABEZADOS
    # ---------------------------------------------

    encabezados = []
    contador = {}

    for encabezado in encabezados_originales:

        if encabezado is None:
            encabezado = ""

        encabezado = str(encabezado).strip()

        # Evitar columnas duplicadas.
        if encabezado in contador:

            contador[encabezado] += 1

            encabezado_final = (
                f"{encabezado}."
                f"{contador[encabezado]}"
            )

        else:
            contador[encabezado] = 0
            encabezado_final = encabezado

        encabezados.append(
            encabezado_final
        )

    datos = filas[
        fila_header + 1:
    ]

    df = pd.DataFrame(
        datos,
        columns=encabezados,
    )

    return df

def limpiar_codigo(valor):
    if pd.isna(valor):
        return ""

    valor = str(valor).strip()

    if valor.lower() in ["nan", "none", "null", ""]:
        return ""

    try:
        valor = str(int(float(valor)))
    except:
        valor = valor.replace(".0", "")

    return valor.lstrip("0")


def limpiar_precio(valor):
    print("ENTRADA:", repr(valor))
    if pd.isna(valor):
        return np.nan

    valor = str(valor).strip()

    if valor == "" or valor.lower() in ["nan", "none", "null"]:
        return np.nan

    # Formato inglés: 2,700.00
    if "," in valor and "." in valor:
        if valor.rfind(".") > valor.rfind(","):
            valor = valor.replace(",", "")
        else:
            valor = valor.replace(".", "").replace(",", ".")

    # Formato argentino: 2700,00
    elif "," in valor:
        valor = valor.replace(",", ".")

    try:
        return float(valor)
    except:
        return np.nan


def limpiar_ean(valor):
    if pd.isna(valor):
        return ""

    valor = str(valor).strip()

    if valor.lower() in ["nan", "none", "null", ""]:
        return ""

    valor = valor.replace(" ", "").replace("-", "")

    try:
        if "e" in valor.lower():
            valor = str(int(float(valor)))
    except:
        pass

    if re.fullmatch(r"\d+\.0", valor):
        valor = valor[:-2]

    valor = re.sub(r"\D", "", valor)

    return valor.strip()


# ---------------- EAN MAP ----------------

def cargar_material_map():
    try:
        material_df = pd.read_excel(
            RUTA_MATERIAL,
            sheet_name="Hoja2",
            dtype=str,
            header=1
        )

        material_df.columns = material_df.columns.str.strip().str.lower()

        material_df["material"] = material_df["material"].apply(limpiar_codigo)
        material_df["scaner"] = material_df["scaner"].apply(limpiar_ean)

        material_df = material_df[
            (material_df["material"] != "") &
            (material_df["scaner"] != "")
        ]

        #print("MATERIAL_MAP registros:", len(material_df))

        return dict(zip(material_df["material"], material_df["scaner"]))

    except Exception as e:
        #print("Error cargando archivo EAN:", e)
        return {}



# mappeo de departamentos

def cargar_departamento_map():
    try:
        material_df = pd.read_excel(
            RUTA_MATERIAL,
            sheet_name="Hoja2",
            dtype=str,
            header=1
        )

        material_df.columns = material_df.columns.str.strip().str.lower()

        material_df["scaner"] = material_df["scaner"].apply(limpiar_ean)
        material_df["departamento"] = material_df["departamento"].astype(str).str.strip()

        material_df = material_df[material_df["scaner"] != ""]

        #print("DEPARTAMENTO_MAP registros:", len(material_df))

        return dict(zip(material_df["scaner"], material_df["departamento"]))

    except Exception as e:
        #print("Error cargando departamentos:", e)
        return {}


def cargar_dep_map():
    try:
        material_df = pd.read_excel(
            RUTA_MATERIAL,
            sheet_name="Hoja2",
            dtype=str,
            header=1
        )

        material_df.columns = material_df.columns.str.strip().str.lower()

        material_df["scaner"] = material_df["scaner"].apply(limpiar_ean)
        material_df["dep"] = material_df["dep"].astype(str).str.strip()

        material_df = material_df[material_df["scaner"] != ""]

        #print("DEP_MAP registros:", len(material_df))

        return dict(zip(material_df["scaner"], material_df["dep"]))

    except Exception as e:
        print("Error cargando DEP:", e)
        return {}


MATERIAL_MAP = cargar_material_map()
DEPARTAMENTO_MAP = cargar_departamento_map()
DEP_MAP = cargar_dep_map()

def completar_ean(df):
    if "CODIGO" not in df.columns:
        #print("DEBUG EAN: no existe columna CODIGO")
        return df

    #print("DEBUG columnas antes EAN:", df.columns.tolist())

    codigos_str = df["CODIGO"].apply(limpiar_codigo)
    mapped = codigos_str.map(MATERIAL_MAP).fillna("")

    #print("DEBUG MATERIAL_MAP size:", len(MATERIAL_MAP))
    #print("DEBUG CODIGOS archivo:", codigos_str.head(10).tolist())
    #print("DEBUG EAN desde MATERIAL_MAP:", mapped.head(10).tolist())

    if "EAN" in df.columns:
        #print("DEBUG EAN original archivo:", df["EAN"].head(10).tolist())

        df["EAN"] = df["EAN"].apply(limpiar_ean)

        #print("DEBUG EAN limpio archivo:", df["EAN"].head(10).tolist())

        df["EAN"] = df["EAN"].mask(df["EAN"] == "", mapped)
    else:
        #print("DEBUG no vino columna EAN, se inserta desde MATERIAL_MAP")
        df.insert(df.columns.get_loc("CODIGO") + 1, "EAN", mapped)

    #print("DEBUG EAN final:", df["EAN"].head(10).tolist())
    #print("DEBUG EAN vacios:", (df["EAN"] == "").sum())

    return df



def completar_departamento(df):

    if "EAN" not in df.columns:
        return df

    ean_str = df["EAN"].apply(limpiar_ean)

    departamento_map_normalized = {}

    for k, v in DEPARTAMENTO_MAP.items():

        key = (
            str(k)
            .strip()
            .replace(".0", "")
        )

        departamento_map_normalized[key] = v

    mapped = ean_str.map(
        departamento_map_normalized
    ).fillna("")

    if "departamento" in df.columns:

        df["departamento"] = df["departamento"].mask(
            df["departamento"].astype(str).str.strip() == "",
            mapped
        ).fillna(mapped)

    else:

        df.insert(
            df.columns.get_loc("EAN") + 1,
            "departamento",
            mapped
        )

    return df


def completar_dep(df):

    if "EAN" not in df.columns:
        return df

    ean_str = df["EAN"].apply(limpiar_ean)

    dep_map_normalized = {}

    for k, v in DEP_MAP.items():

        key = (
            str(k)
            .strip()
            .replace(".0", "")
        )

        dep_map_normalized[key] = v

    mapped = ean_str.map(
        dep_map_normalized
    ).fillna("")

    if "dep" in df.columns:

        df["dep"] = df["dep"].mask(
            df["dep"].astype(str).str.strip() == "",
            mapped
        ).fillna(mapped)

    else:

        df.insert(
            df.columns.get_loc("departamento") + 1,
            "dep",
            mapped
        )

    return df


# ---------------- UTIL ----------------

def normalizar_texto(texto):
    if not texto:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = texto.encode("ascii", "ignore").decode("utf-8")
    return texto

# ---------------- PROCESAMIENTO COMUN ----------------


COLUMNAS_EDITABLES_PREVIEW = {
    "DESCRIPCION",
    "Normal",
    "Oferta",
    "cenefa",
    "desde",
    "hasta",
    "sucursales",
}


def normalizar_cambios_preview(cambios_json):

    try:
        cambios = json.loads(
            cambios_json or "[]"
        )

    except json.JSONDecodeError as error:
        raise ValueError(
            "No se pudieron interpretar "
            "las modificaciones."
        ) from error

    if not isinstance(cambios, list):
        raise ValueError(
            "El formato de las modificaciones "
            "no es válido."
        )

    if len(cambios) > 10000:
        raise ValueError(
            "Se superó la cantidad máxima "
            "de modificaciones permitidas."
        )

    return cambios


def aplicar_cambios_preview(
    df,
    cambios,
):

    df = df.copy().reset_index(
        drop=True
    )

    for cambio in cambios:

        if not isinstance(cambio, dict):
            raise ValueError(
                "Se recibió una modificación "
                "inválida."
            )

        fila = cambio.get("fila")
        columna = cambio.get("columna")
        valor = cambio.get("valor")

        # ====================================================
        # VALIDAR FILA
        # ====================================================

        if not isinstance(fila, int):
            raise ValueError(
                "Número de fila inválido."
            )

        if fila < 0 or fila >= len(df):
            raise ValueError(
                f"La fila {fila + 1} "
                "no existe."
            )

        # ====================================================
        # VALIDAR COLUMNA EDITABLE
        # ====================================================

        if columna not in COLUMNAS_EDITABLES_PREVIEW:
            raise ValueError(
                f"No está permitido modificar "
                f"{columna}."
            )

        if columna not in df.columns:
            raise ValueError(
                f"La columna {columna} "
                "no existe."
            )

        # ====================================================
        # PRECIOS
        # ====================================================

        if columna in {
            "Normal",
            "Oferta",
        }:

            valor = limpiar_precio(
                valor
            )

            if pd.isna(valor):
                raise ValueError(
                    f"Precio inválido en "
                    f"fila {fila + 1}, "
                    f"columna {columna}."
                )

            if valor < 0:
                raise ValueError(
                    f"El precio de la fila "
                    f"{fila + 1} no puede "
                    "ser negativo."
                )

            valor = (
                np.floor(valor * 100)
                / 100
            )

        # ====================================================
        # DESCRIPCION / CENEFA
        # ====================================================

        elif columna in {
            "DESCRIPCION",
            "cenefa",
        }:

            valor = str(
                valor or ""
            ).strip()

            if len(valor) > 250:
                raise ValueError(
                    f"El campo {columna} "
                    f"de la fila {fila + 1} "
                    "supera los 250 caracteres."
                )

            if (
                columna == "cenefa"
                and not valor
            ):
                valor = "OFERTA"

        # ====================================================
        # FECHAS
        # ====================================================

        elif columna in {
            "desde",
            "hasta",
        }:

            valor = str(
                valor or ""
            ).strip()

            if not valor:
                raise ValueError(
                    f"La fecha {columna} "
                    f"de la fila {fila + 1} "
                    "no puede estar vacía."
                )

            try:
                datetime.strptime(
                    valor,
                    "%Y-%m-%d",
                )

            except ValueError as error:
                raise ValueError(
                    f"La fecha {columna} "
                    f"de la fila {fila + 1} "
                    "no tiene un formato válido. "
                    "Debe ser YYYY-MM-DD."
                ) from error

        # ====================================================
        # SUCURSALES
        # ====================================================

        elif columna == "sucursales":

            valor = str(
                valor or ""
            ).strip().upper()

            if not valor:
                raise ValueError(
                    f"Las sucursales de la fila "
                    f"{fila + 1} no pueden "
                    "estar vacías."
                )

            # Limpiamos espacios:
            # "CO01, CO02" -> "CO01,CO02"
            sucursales = [
                sucursal.strip()
                for sucursal
                in valor.split(",")
                if sucursal.strip()
            ]

            if not sucursales:
                raise ValueError(
                    f"No se encontraron sucursales "
                    f"válidas en la fila {fila + 1}."
                )

            # Validar formato COxx o MAxx
            for sucursal in sucursales:

                if not re.fullmatch(
                    r"(CO|MA)[0-9]{2}",
                    sucursal,
                ):
                    raise ValueError(
                        f"La sucursal '{sucursal}' "
                        f"de la fila {fila + 1} "
                        "no tiene un formato válido."
                    )

            # Evitar duplicados manteniendo orden.
            sucursales = list(
                dict.fromkeys(
                    sucursales
                )
            )

            valor = ",".join(
                sucursales
            )

        # ====================================================
        # APLICAR CAMBIO
        # ====================================================

        df.at[
            fila,
            columna,
        ] = valor

    # ========================================================
    # VALIDAR COHERENCIA DE FECHAS
    # ========================================================

    for indice, row in df.iterrows():

        desde = str(
            row.get("desde") or ""
        ).strip()

        hasta = str(
            row.get("hasta") or ""
        ).strip()

        if desde and hasta:

            try:
                fecha_desde = datetime.strptime(
                    desde,
                    "%Y-%m-%d",
                ).date()

                fecha_hasta = datetime.strptime(
                    hasta,
                    "%Y-%m-%d",
                ).date()

            except ValueError as error:
                raise ValueError(
                    f"Las fechas de la fila "
                    f"{indice + 1} "
                    "no son válidas."
                ) from error

            if fecha_desde > fecha_hasta:
                raise ValueError(
                    f"En la fila {indice + 1}, "
                    "la fecha Desde no puede "
                    "ser posterior a Hasta."
                )

    return df



def procesar_archivo_cenefas(archivo, tipo, fecha_desde, fecha_hasta):
    preview = None
    mensaje_error = None
    total_registros = 0
    df = None

    if archivo and archivo.filename != "":

        excel_file = pd.ExcelFile(archivo)

        hoja_objetivo = None
        for hoja in excel_file.sheet_names:
            if normalizar_texto(hoja).strip() in ["cenefa", "cenefas"]:
                hoja_objetivo = hoja
                break

        if hoja_objetivo is None:
            hoja_objetivo = excel_file.sheet_names[0]

        df_temp = pd.read_excel(
            excel_file,
            sheet_name=hoja_objetivo,
            header=None
        )

        fila_header = None

        for i, row in df_temp.iterrows():
            valores = [normalizar_texto(x) for x in row.values]

            if "codigo" in valores:
                fila_header = i
                break

        if fila_header is None:
            mensaje_error = "No se encontró la fila de encabezados (CODIGO)."
            return None, None, mensaje_error, 0

        #df = pd.read_excel(
        #    excel_file,
        #    sheet_name=hoja_objetivo,
        #    header=fila_header,
        #    dtype=str
        #)

        df = crear_dataframe_solo_valores(
            archivo=archivo,
            hoja_objetivo=hoja_objetivo,
            fila_header=fila_header,
        )

        df.columns = [
            normalizar_texto(col).strip()
            for col in df.columns
        ]

        #columnas_cenefa = ["cenefa", "cenefas", "promo", "promos"]

        #col_cenefa_detectada = None

        #for col in df.columns:
        #    if normalizar_texto(col) in columnas_cenefa:
        #        col_cenefa_detectada = col
        #        break

        #if col_cenefa_detectada:
        #    df = df.rename(columns={col_cenefa_detectada: "cenefa"})
        #else:
        #    df["cenefa"] = "OFERTA"


        # -----------------------------------------------------

        column_mapping = {}

        for header in HEADERS:

            header_norm = normalizar_texto(header)

            posibles = ALIAS.get(header, [])
            posibles_norm = [
                normalizar_texto(p)
                for p in posibles
            ]

            for col in df.columns:

                if col == header_norm or col in posibles_norm:
                    column_mapping[col] = header
                    break

        if not column_mapping:
            mensaje_error = "No se encontraron columnas válidas."
            return None, None, mensaje_error, 0

        df = df.rename(columns=column_mapping)

        if "cenefa" not in df.columns:
            df["cenefa"] = "OFERTA"
        else:
            df["cenefa"] = (
                df["cenefa"]
                .replace(r"^\s*$", pd.NA, regex=True)
                .replace(["nan", "NaN", "None", "none"], pd.NA)
                .ffill()
                .fillna("OFERTA")
            )

        if "sucursales" in df.columns:
            df["sucursales"] = df["sucursales"].ffill()

        #if "cenefa" in df.columns:
        #    df["cenefa"] = df["cenefa"].ffill()

        if "Oferta" in df.columns:
            df["Oferta"] = df["Oferta"].ffill()

        if "Normal" in df.columns:
            df["Normal"] = df["Normal"].ffill()

        #for col in ["sucursales", "Oferta", "Normal"]:
        #    if col in df.columns:
        #        df[col] = (
        #            df[col]
        #            .replace(r"^\s*$", pd.NA, regex=True)
        #            .replace(["nan", "NaN", "None", "none"], pd.NA)
        #            .ffill()
        #        )

        #if "cenefa" in df.columns:
        #    df["cenefa"] = (
        #        df["cenefa"]
        #        .replace(r"^\s*$", pd.NA, regex=True)
        #        .replace(["nan", "NaN", "None", "none"], pd.NA)
        #        .ffill()
        #    )
        #else:
        #    df["cenefa"] = "OFERTA"

        if tipo in SUCURSAL_MAP:

            clave_total = (
                "Total Empresa - Mayorista"
                if tipo == "mayorista"
                else "Total-empresa-minorista"
            )

            if "sucursales" not in df.columns:
                df["sucursales"] = ""

            #def generar_codigos(valor):

            #    if pd.isna(valor) or str(valor).strip() == "":
            #        return SUCURSAL_MAP[tipo].get(clave_total, "")

            #    provincia = normalizar_texto(valor)

            #    return SUCURSAL_MAP[tipo].get(
            #        provincia,
            #        SUCURSAL_MAP[tipo].get(clave_total, "")
            #    )


            def generar_codigos(valor):

                if pd.isna(valor) or str(valor).strip() == "":
                    return SUCURSAL_MAP[tipo].get(clave_total, "")

                valor_original = str(valor).strip().upper()
                valor_norm = normalizar_texto(valor)

                if valor_norm in ["", "interna"]:
                    return SUCURSAL_MAP[tipo].get(clave_total, "")

                # Si ya viene un código de sucursal tipo CO24 o MA02
                if re.fullmatch(r"(CO\d{2}|MA\d{2})", valor_original):
                    return valor_original

                return SUCURSAL_MAP[tipo].get(
                    valor_norm,
                    SUCURSAL_MAP[tipo].get(clave_total, "")
                )


            df["sucursales"] = df["sucursales"].apply(
                generar_codigos
            )

        columnas_validas = [
            col for col in df.columns
            if col in HEADERS
        ]

        df = df[columnas_validas]

        df = df.replace(r'^\s*$', pd.NA, regex=True)
        df = df.dropna(how="all")

        if "CODIGO" in df.columns:
            df["CODIGO"] = pd.to_numeric(
                df["CODIGO"],
                errors="coerce"
            )

            df = df.dropna(subset=["CODIGO"])
            df["CODIGO"] = df["CODIGO"].astype(int)

        df = completar_ean(df)
        df = completar_departamento(df)
        df = completar_dep(df)

        columnas = list(df.columns)

        if "CODIGO" in columnas and "EAN" in columnas:

            columnas.remove("EAN")

            index_codigo = columnas.index("CODIGO")

            columnas.insert(index_codigo + 1, "EAN")

            df = df[columnas]

    #    def limpiar_precio(valor):

    #        if pd.isna(valor):
    #            return np.nan

    #        valor = str(valor).strip()

    #        if "," in valor and "." in valor:
    #            valor = valor.replace(",", "")

    #        elif "," in valor:
    #            valor = valor.replace(",", ".")

    #        try:
    #            return float(valor)

    #        except:
    #            return np.nan
    if df is None:
        return None, None, mensaje_error, 0

    if "Oferta" in df.columns:
        print("VALORES ORIGINALES OFERTA")
        print(df["Oferta"].head(20).tolist())
        df["Oferta"] = df["Oferta"].apply(limpiar_precio)
        df["Oferta"] = np.floor(df["Oferta"] * 100) / 100

    if "Normal" in df.columns:
        df["Normal"] = df["Normal"].apply(limpiar_precio)
        df["Normal"] = np.floor(df["Normal"] * 100) / 100

    if fecha_desde:
        df["desde"] = fecha_desde

    if fecha_hasta:
        df["hasta"] = fecha_hasta

    df = df.reset_index(drop=True)

    total_registros = len(df)

    #df = df.fillna("")
    df_preview = df.fillna("")

    preview = df_preview.to_html(
            classes="table table-striped table-bordered",
            index=False
    )

    return df, preview, mensaje_error, total_registros




#funcion para formatear precios en bd
#def limpiar_precio(valor):

#    if pd.isna(valor):
#        return np.nan

#    valor = str(valor).strip()

#    if valor == "" or valor.lower() in ["nan", "none", "null"]:
#        return np.nan

    # 5.700,00 -> 5700.00
#    if "," in valor and "." in valor:
#        valor = valor.replace(".", "").replace(",", ".")

    # 5700,00 -> 5700.00
#    elif "," in valor:
#        valor = valor.replace(",", ".")

#    try:
#        return float(valor)
#    except:
#        return np.nan



#funcion para formatear precios en vistas
def formatear_precio_arg(valor):

    if valor is None or valor == "":
        return ""

    try:
        valor = float(valor)
        return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return valor



#funcion para insertar registros vacios a bd
def valor_db(valor):

    if pd.isna(valor):
        return None
    return valor








# Controlar cenefas repetidas

def existen_cenefas_repetidas(df, tipo_cenefa):
    #conn = sqlite3.connect(DB_PATH)
    conn = get_db_connection()
    cursor = conn.cursor()

    repetidos = []

    for _, row in df.iterrows():
        cursor.execute("""
            SELECT COUNT(*)
            FROM cenefas
            WHERE Codigo = %s
              AND tipo_cenefa = %s
              AND desde = %s
              AND hasta = %s
        """, (
            row.get("CODIGO"),
            tipo_cenefa,
            row.get("desde"),
            row.get("hasta")
        ))

        if cursor.fetchone()[0] > 0:
            repetidos.append(row.get("CODIGO"))

    conn.close()
    return repetidos

def guardar_cenefas_en_db(df, tipo_cenefa, usuario="sistema", lote_carga=None, sobrescribir=False):
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        if lote_carga is None:
            lote_carga = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + str(uuid.uuid4())[:8]

        fecha_carga = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        col_codigo = None
        col_cenefa = None
        col_descripcion = None

        for col in df.columns:
            col_norm = str(col).strip().lower()

            if col_norm == "codigo":
                col_codigo = col

            elif col_norm in ["cenefa", "cenefas"]:
                col_cenefa = col

            elif col_norm in ["descripcion", "descripción"]:
                col_descripcion = col

        if not col_codigo:
            raise ValueError("No se encontró columna CODIGO para guardar cenefas.")

        if not col_cenefa:
            df["cenefa"] = ""
            col_cenefa = "cenefa"

        df = df[
            df[col_codigo].notna()
            & (df[col_codigo].astype(str).str.strip() != "")
            & (df[col_codigo].astype(str).str.strip().str.lower() != "codigo")
        ].copy()

        columnas_control = {
            col_codigo: "codigo",
            col_cenefa: "cenefa",
        }

        if col_descripcion:
            columnas_control[col_descripcion] = "descripcion"

        for col, valor_header in columnas_control.items():
            if col in df.columns:
                df = df[
                    df[col].isna()
                    | (df[col].astype(str).str.strip().str.lower() != valor_header)
                ].copy()

        if df.empty:
            raise ValueError(
                f"No hay registros válidos para guardar en cenefas ({tipo_cenefa})"
            )

#        if sobrescribir:
#            for _, row in df.iterrows():
#                desde = row.get("desde")
#                hasta = row.get("hasta")

#                cursor.execute("""
#                    DELETE FROM cenefas
#                    WHERE Codigo = %s
#                      AND tipo_cenefa = %s
#                      AND desde = %s
#                      AND hasta = %s
#                """, (
#                    valor_db(row.get(col_codigo)),
#                    valor_db(tipo_cenefa),
#                    valor_db(desde),
#                    valor_db(hasta)
#                ))

        if sobrescribir:
            for _, row in df.iterrows():
                desde = row.get("desde")
                hasta = row.get("hasta")

                cursor.execute(
                    """
                    DELETE FROM cenefas
                    WHERE Codigo = %s
                    AND tipo_cenefa = %s
                    AND desde = %s
                    AND hasta = %s
                    AND sucursales = %s
                    """,
                    (
                        valor_db(row.get(col_codigo)),
                        valor_db(tipo_cenefa),
                        valor_db(desde),
                        valor_db(hasta),
                        valor_db(row.get("sucursales")),
                    ),
                )


        for _, row in df.iterrows():
            desde = row.get("desde")
            hasta = row.get("hasta")

            cursor.execute("""
                INSERT INTO cenefas
                (
                    Codigo, ean, dep, departamento, descripcion, Normal, Oferta, cenefa,
                    desde, hasta, sucursales, tipo_cenefa,
                    fecha_carga, lote_carga, usuario_carga
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                valor_db(row.get(col_codigo)),
                valor_db(row.get("EAN")),
                valor_db(row.get("dep")),
                valor_db(row.get("departamento")),
                valor_db(row.get(col_descripcion) if col_descripcion else row.get("DESCRIPCION")),
                valor_db(row.get("Normal")),
                valor_db(row.get("Oferta")),
                valor_db(row.get(col_cenefa)),
                valor_db(desde),
                valor_db(hasta),
                valor_db(row.get("sucursales")),
                valor_db(tipo_cenefa),
                valor_db(fecha_carga),
                valor_db(lote_carga),
                valor_db(usuario)
                ))

        conn.commit()
        return lote_carga, fecha_carga

    except Exception:
        conn.rollback()
        raise

    finally:
        cursor.close()
        conn.close()








# ---------------- ROUTES ----------------

@compras_bp.route("/")
def dashboard():
    return render_template("compras.html")

def _folder_base(tipo, template_name):
    preview = None
    mensaje_error = None
    total_registros = 0

    fecha_desde = (
        request.form.get("fecha_desde")
        or request.args.get("fecha_desde")
        or ""
    )
    fecha_hasta = (
        request.form.get("fecha_hasta")
        or request.args.get("fecha_hasta")
        or ""
    )

    if request.method == "POST":
        archivo = request.files.get("archivo")
        usuario = session.get("usuario_nombre", "desconocido")

        try:
            df, _, mensaje_error, total_registros = procesar_archivo_cenefas(
                archivo=archivo,
                tipo=tipo,
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
            )

            if df is None:
                raise ValueError(
                    mensaje_error or "No se pudo procesar el archivo."
                )

            if df.empty:
                raise ValueError("El archivo no contiene registros válidos.")

            df = df.reset_index(drop=True)
            lote_id = uuid.uuid4().hex

            # Solo se guarda temporalmente. La BD se modifica al transmitir.
            guardar_temporal(lote_id, df)

            session["folder_lote_id"] = lote_id
            session["folder_preview_tipo"] = tipo
            session["folder_preview_fecha_desde"] = fecha_desde
            session["folder_preview_fecha_hasta"] = fecha_hasta
            session["folder_preview_template"] = template_name

            preview = df.fillna("").to_dict(orient="records")
            total_registros = len(df)

            guardar_log_compras(
                usuario=usuario,
                nivel="INFO",
                origen="backend",
                modulo="folder",
                accion=f"Previsualizar folder {tipo}",
                archivo=archivo.filename if archivo else None,
                detalle="Archivo procesado. Pendiente de transmisión.",
                estado="exitoso",
                total_registros=total_registros,
            )

        except Exception as error:
            mensaje_error = str(error)

            guardar_log_compras(
                usuario=usuario,
                nivel="ERROR",
                origen="backend",
                modulo="folder",
                accion=f"Error previsualizando folder {tipo}",
                archivo=archivo.filename if archivo else None,
                detalle=str(error),
                estado="fallido",
                total_registros=0,
            )

    return render_template(
        template_name,
        preview=preview,
        tipo=tipo,
        mensaje_error=mensaje_error,
        total_registros=total_registros,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        columnas_editables=COLUMNAS_EDITABLES_PREVIEW,
    )


@compras_bp.route("/folder/mayorista", methods=["GET", "POST"])
@login_requerido("compras")
def folder_mayorista():
    return _folder_base("mayorista", "folder-mayorista.html")


@compras_bp.route("/folder/minorista", methods=["GET", "POST"])
@login_requerido("compras")
def folder_minorista():
    return _folder_base("minorista", "folder-minorista.html")


@compras_bp.route("/folder/transmitir", methods=["POST"])
@login_requerido("compras")
def transmitir_folder():
    lote_id = session.get("folder_lote_id")
    tipo = session.get("folder_preview_tipo")
    fecha_desde = session.get("folder_preview_fecha_desde", "")
    fecha_hasta = session.get("folder_preview_fecha_hasta", "")
    template_name = session.get("folder_preview_template")
    usuario = session.get("usuario_nombre", "desconocido")

    if not lote_id or not tipo:
        return "No existe una previsualización para transmitir.", 400

    if not template_name:
        template_name = (
            "folder-mayorista.html"
            if tipo == "mayorista"
            else "folder-minorista.html"
        )

    df = recuperar_temporal(lote_id)

    if df is None:
        return render_template(
            template_name,
            preview=None,
            tipo=tipo,
            mensaje_error=(
                "La previsualización venció o fue eliminada. "
                "Debe volver a cargar el archivo."
            ),
            total_registros=0,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            columnas_editables=COLUMNAS_EDITABLES_PREVIEW,
        ), 400

    try:
        cambios_json = request.form.get("cambios", "[]")
        sobrescribir = request.form.get("sobrescribir") == "1"

        cambios = normalizar_cambios_preview(cambios_json)
        df = aplicar_cambios_preview(df, cambios)

        # Guardamos el estado editado para no perder cambios si hay que
        # confirmar una sobrescritura en un segundo envío.
        guardar_temporal(lote_id, df)

        repetidos = existen_cenefas_repetidas(df, tipo)

        if repetidos and not sobrescribir:
            return render_template(
                template_name,
                preview=df.fillna("").to_dict(orient="records"),
                tipo=tipo,
                mensaje_error=(
                    f"Ya existen {len(repetidos)} registros para este período. "
                    "Confirme el reemplazo para sobrescribirlos."
                ),
                total_registros=len(df),
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
                requiere_sobrescribir=True,
                columnas_editables=COLUMNAS_EDITABLES_PREVIEW,
            )

        lote_carga, fecha_carga = guardar_cenefas_en_db(
            df=df,
            tipo_cenefa=tipo,
            usuario=usuario,
            sobrescribir=sobrescribir,
        )

        redis_client.delete(lote_id)

        session.pop("folder_lote_id", None)
        session.pop("folder_preview_tipo", None)
        session.pop("folder_preview_fecha_desde", None)
        session.pop("folder_preview_fecha_hasta", None)
        session.pop("folder_preview_template", None)

        guardar_log_compras(
            usuario=usuario,
            nivel="INFO",
            origen="backend",
            modulo="folder",
            accion=f"Transmitir folder {tipo}",
            detalle=f"Folder transmitido correctamente. Lote: {lote_carga}",
            estado="exitoso",
            total_registros=len(df),
        )

        return render_template(
            template_name,
            preview=None,
            tipo=tipo,
            mensaje_error=None,
            mensaje_exito="Folder transmitido correctamente.",
            total_registros=0,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            lote_carga=lote_carga,
            fecha_carga=fecha_carga,
            columnas_editables=COLUMNAS_EDITABLES_PREVIEW,
        )

    except psycopg2.Error as error:
        guardar_log_compras(
            usuario=usuario,
            nivel="CRITICAL",
            origen="base_datos",
            modulo="folder",
            accion=f"Error transmitiendo folder {tipo}",
            detalle=str(error),
            estado="fallido",
            total_registros=len(df) if df is not None else 0,
        )

        return render_template(
            template_name,
            preview=df.fillna("").to_dict(orient="records"),
            tipo=tipo,
            mensaje_error=f"Error de base de datos: {error}",
            total_registros=len(df),
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            columnas_editables=COLUMNAS_EDITABLES_PREVIEW,
        ), 500

    except Exception as error:
        guardar_log_compras(
            usuario=usuario,
            nivel="ERROR",
            origen="backend",
            modulo="folder",
            accion=f"Excepción transmitiendo folder {tipo}",
            detalle=str(error),
            estado="fallido",
            total_registros=len(df) if df is not None else 0,
        )

        return render_template(
            template_name,
            preview=df.fillna("").to_dict(orient="records"),
            tipo=tipo,
            mensaje_error=f"Error transmitiendo el folder: {error}",
            total_registros=len(df),
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            columnas_editables=COLUMNAS_EDITABLES_PREVIEW,
        ), 500


@compras_bp.route("/cenefas", methods=["GET", "POST"])
def cenefas():
    tipo = request.args.get("tipo") or "minorista"
    return render_template("cenefas.html", tipo=tipo)


@compras_bp.route("/ofertas/<modo>", methods=["GET", "POST"])
@login_requerido("compras")
def ofertas(modo):
    modos_validos = {
        "competencia": "Oferta por Competencia",
        "interna": "Oferta Interna",
        "vencimientos": "Oferta por Vencimientos",
    }

    if modo not in modos_validos:
        return "Modo no válido", 404

    preview = None
    mensaje_error = None
    total_registros = 0
    fecha_desde = ""
    fecha_hasta = ""
    tipo = "mayorista"

    if request.method == "POST":
        archivo = request.files.get("archivo")
        fecha_desde = request.form.get("fecha_desde", "").strip()
        fecha_hasta = request.form.get("fecha_hasta", "").strip()
        tipo = request.form.get("tipo", "mayorista")
        usuario = session.get("usuario_nombre", "desconocido")

        try:
            df, _, mensaje_error, total_registros = procesar_archivo_cenefas(
                archivo=archivo,
                tipo=tipo,
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
            )

            if df is None:
                raise ValueError(
                    mensaje_error or "No se pudo procesar el archivo."
                )

            if df.empty:
                raise ValueError("El archivo no contiene registros válidos.")

            df = df.reset_index(drop=True)
            lote_id = uuid.uuid4().hex

            guardar_temporal(lote_id, df)

            session["ofertas_lote_id"] = lote_id
            session["ofertas_preview_modo"] = modo
            session["ofertas_preview_tipo"] = tipo
            session["ofertas_preview_fecha_desde"] = fecha_desde
            session["ofertas_preview_fecha_hasta"] = fecha_hasta

            preview = df.fillna("").to_dict(orient="records")
            total_registros = len(df)

            guardar_log_compras(
                usuario=usuario,
                nivel="INFO",
                origen="backend",
                modulo="ofertas",
                accion=f"Previsualizar oferta {modo}",
                archivo=archivo.filename if archivo else None,
                detalle="Archivo procesado. Pendiente de transmisión.",
                estado="exitoso",
                total_registros=total_registros,
            )

        except Exception as error:
            mensaje_error = str(error)

            guardar_log_compras(
                usuario=usuario,
                nivel="ERROR",
                origen="backend",
                modulo="ofertas",
                accion=f"Error previsualizando oferta {modo}",
                archivo=archivo.filename if archivo else None,
                detalle=str(error),
                estado="fallido",
                total_registros=0,
            )

    return render_template(
        "ofertas.html",
        modo=modo,
        titulo_vista=modos_validos[modo],
        preview=preview,
        tipo=tipo,
        mensaje_error=mensaje_error,
        total_registros=total_registros,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        columnas_editables=COLUMNAS_EDITABLES_PREVIEW,
    )


@compras_bp.route("/transmitir_ofertas", methods=["POST"])
@login_requerido("compras")
def transmitir_ofertas():
    lote_id = session.get("ofertas_lote_id")
    modo = session.get("ofertas_preview_modo")
    tipo = session.get("ofertas_preview_tipo", "mayorista")
    fecha_desde = session.get("ofertas_preview_fecha_desde", "")
    fecha_hasta = session.get("ofertas_preview_fecha_hasta", "")
    usuario = session.get("usuario_nombre", "desconocido")

    modos_validos = {
        "competencia": "Oferta por Competencia",
        "interna": "Oferta Interna",
        "vencimientos": "Oferta por Vencimientos",
    }

    if not lote_id or not modo or modo not in modos_validos:
        return "No hay datos para transmitir.", 400

    df = recuperar_temporal(lote_id)

    if df is None:
        return render_template(
            "ofertas.html",
            modo=modo,
            titulo_vista=modos_validos[modo],
            preview=None,
            tipo=tipo,
            mensaje_error=(
                "La previsualización venció o fue eliminada. "
                "Debe volver a cargar el archivo."
            ),
            total_registros=0,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            columnas_editables=COLUMNAS_EDITABLES_PREVIEW,
        ), 400

    try:
        cambios_json = request.form.get("cambios", "[]")
        sobrescribir = request.form.get("sobrescribir") == "1"

        cambios = normalizar_cambios_preview(cambios_json)
        df = aplicar_cambios_preview(df, cambios)

        # Persistimos las ediciones en Redis para conservarlas si aparece
        # la confirmación de sobrescritura.
        guardar_temporal(lote_id, df)

        repetidos = existen_cenefas_repetidas(df, modo)

        if repetidos and not sobrescribir:
            return render_template(
                "ofertas.html",
                modo=modo,
                titulo_vista=modos_validos[modo],
                preview=df.fillna("").to_dict(orient="records"),
                tipo=tipo,
                mensaje_error=(
                    f"Ya existen {len(repetidos)} registros para este período. "
                    "Confirme el reemplazo para sobrescribirlos."
                ),
                total_registros=len(df),
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
                requiere_sobrescribir=True,
                columnas_editables=COLUMNAS_EDITABLES_PREVIEW,
            )

        lote_carga, fecha_carga = guardar_cenefas_en_db(
            df=df,
            tipo_cenefa=modo,
            usuario=usuario,
            sobrescribir=sobrescribir,
        )

        redis_client.delete(lote_id)

        session.pop("ofertas_lote_id", None)
        session.pop("ofertas_preview_modo", None)
        session.pop("ofertas_preview_tipo", None)
        session.pop("ofertas_preview_fecha_desde", None)
        session.pop("ofertas_preview_fecha_hasta", None)

        guardar_log_compras(
            usuario=usuario,
            nivel="INFO",
            origen="backend",
            modulo="transmitir",
            accion=f"Transmitir oferta {modo}",
            detalle=f"Datos transmitidos correctamente. Lote: {lote_carga}",
            estado="exitoso",
            total_registros=len(df),
        )

        return render_template(
            "ofertas.html",
            modo=modo,
            titulo_vista=modos_validos[modo],
            preview=None,
            tipo=tipo,
            mensaje_error=None,
            mensaje_exito="Datos transmitidos correctamente.",
            total_registros=0,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            lote_carga=lote_carga,
            fecha_carga=fecha_carga,
            columnas_editables=COLUMNAS_EDITABLES_PREVIEW,
        )

    except psycopg2.Error as error:
        guardar_log_compras(
            usuario=usuario,
            nivel="CRITICAL",
            origen="base_datos",
            modulo="transmitir",
            accion=f"Error transmitiendo oferta {modo}",
            detalle=str(error),
            estado="fallido",
            total_registros=len(df) if df is not None else 0,
        )

        return render_template(
            "ofertas.html",
            modo=modo,
            titulo_vista=modos_validos[modo],
            preview=df.fillna("").to_dict(orient="records"),
            tipo=tipo,
            mensaje_error=f"Error de base de datos: {error}",
            total_registros=len(df),
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            columnas_editables=COLUMNAS_EDITABLES_PREVIEW,
        ), 500

    except Exception as error:
        guardar_log_compras(
            usuario=usuario,
            nivel="ERROR",
            origen="backend",
            modulo="transmitir",
            accion=f"Excepción transmitiendo oferta {modo}",
            detalle=str(error),
            estado="fallido",
            total_registros=len(df) if df is not None else 0,
        )

        return render_template(
            "ofertas.html",
            modo=modo,
            titulo_vista=modos_validos[modo],
            preview=df.fillna("").to_dict(orient="records"),
            tipo=tipo,
            mensaje_error=f"Error transmitiendo la oferta: {error}",
            total_registros=len(df),
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            columnas_editables=COLUMNAS_EDITABLES_PREVIEW,
        ), 500


@compras_bp.route("/sucursal")
@login_requerido("sucursal")
def sucursal():
    from datetime import datetime

    def convertir_fecha(valor):
        if not valor:
            return None

        valor = str(valor).strip()
        valor = valor.replace("$", "").replace(" ", "")

        for formato in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(valor, formato).date()
            except:
                pass

        return None
        
    def formatear_fecha(valor):
        fecha = convertir_fecha(valor)
        if not fecha:
            return ""
        return fecha.strftime("%d/%m/%Y")

    sucursal_codigo = session.get("usuario_nombre", "").strip().upper()
    tipo = request.args.get("tipo", "minorista")
    #hoy = datetime.now().date() + timedelta(days=2)
    hoy = datetime.now().date()

    #conn = sqlite3.connect(DB_PATH)
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT Codigo, ean,dep, departamento, descripcion, Normal, Oferta, cenefa,
               desde, hasta, sucursales, tipo_cenefa,
               fecha_carga, lote_carga, usuario_carga
        FROM cenefas
        WHERE tipo_cenefa = %s
        ORDER BY fecha_carga DESC, desde DESC
    """, (tipo,))

    rows = cursor.fetchall()
    conn.close()

    filtradas = []

    for r in rows:
        sucursales = str(r[10]).upper().replace(" ", "")
        lista_suc = [s.strip() for s in sucursales.split(",")]

    
        desde = convertir_fecha(r[8])
        hasta = convertir_fecha(r[9])

        if not desde or not hasta:
            print("ERROR fecha:", r[7], r[8])
            continue

        #print("DEBUG desde:", desde, "hasta:", hasta)

        #if sucursal_codigo in lista_suc and desde <= hoy <= hasta:
        inicio_visualizacion = desde - timedelta(days=3)
        if sucursal_codigo in lista_suc and inicio_visualizacion <= hoy <= hasta:

            filtradas.append({
                "Codigo": r[0],
                "ean": r[1],
                "dep": r[2],
                "departamento": r[3],
                "descripcion": r[4],
                #"Normal": r[5],
                #"Oferta": r[6],
                "Normal": formatear_precio_arg(r[5]),
                "Oferta": formatear_precio_arg(r[6]),
                "cenefa": r[7],
                "desde": formatear_fecha(r[8]),
                "hasta": formatear_fecha(r[9]),
                "sucursales": r[10],
                "tipo_cenefa": r[11],
                "fecha_carga": r[12],
                "lote_carga": r[13],
                "usuario_carga": r[14],
                "es_nueva": False
            })

    lote_mas_reciente = None
    ultima_fecha_carga = None
    ultimo_usuario_carga = None

    if filtradas:
        ordenadas_por_carga = sorted(
            [x for x in filtradas if x["fecha_carga"]],
            key=lambda x: x["fecha_carga"],
            reverse=True
        )

        if ordenadas_por_carga:
            lote_mas_reciente = ordenadas_por_carga[0]["lote_carga"]
            ultima_fecha_carga = ordenadas_por_carga[0]["fecha_carga"]
            ultimo_usuario_carga = ordenadas_por_carga[0]["usuario_carga"]

        for item in filtradas:
            if item["lote_carga"] == lote_mas_reciente:
                item["es_nueva"] = True

    return render_template(
        "sucursales.html",
        datos=filtradas,
        tipo=tipo,
        ultima_fecha_carga=ultima_fecha_carga,
        ultimo_usuario_carga=ultimo_usuario_carga
    )


    # --------------- FARMACIA --------------

@compras_bp.route("/farmacia_folder", methods=["GET", "POST"])
@login_requerido("adm-farmacia")
def farmacia_folder():
    preview = None
    mensaje_error = None
    total_registros = 0
    fecha_desde = None
    fecha_hasta = None

    if request.method == "POST":
        archivo = request.files.get("archivo")
        fecha_desde = request.form.get("fecha_desde")
        fecha_hasta = request.form.get("fecha_hasta")

        df, preview, mensaje_error, total_registros = procesar_archivo_cenefas(
            archivo=archivo,
            tipo=None,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta
        )

        if df is not None:
            if "sucursales" not in df.columns:
                df["sucursales"] = ""

            guardar_cenefas_en_db(df, "farmacia")

    return render_template(
        "farmacia_folder.html",
        preview=preview,
        mensaje_error=mensaje_error,
        total_registros=total_registros,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta
    )



# Historial cenefas

@compras_bp.route("/historico")
@login_requerido("sucursal")
def historico_cenefas():
    filtro_codigo = request.args.get("codigo", "").strip()
    filtro_tipo = request.args.get("tipo", "").strip()
    filtro_lote = request.args.get("lote", "").strip()

    #conn = sqlite3.connect(DB_PATH)
    conn = get_db_connection()
    #conn.row_factory = sqlite3.Row
    #cursor = conn.cursor()
    cursor = conn.cursor(cursor_factory=RealDictCursor)

    query = """
        SELECT id, codigo, ean, dep, departamento, descripcion,
            normal, oferta, cenefa,
            desde, hasta, sucursales, tipo_cenefa,
            fecha_carga, lote_carga, usuario_carga
        FROM cenefas
        WHERE 1=1
    """
    params = []

    if filtro_codigo:
        #query += " AND Codigo LIKE %s"
        query += " AND Codigo::text LIKE %s"
        params.append(f"%{filtro_codigo}%")

    if filtro_tipo:
        query += " AND tipo_cenefa = %s"
        params.append(filtro_tipo)

    if filtro_lote:
        query += " AND lote_carga LIKE %s"
        params.append(f"%{filtro_lote}%")

    query += " ORDER BY fecha_carga DESC, id DESC"

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()

    datos = []
    for r in rows:
        datos.append({
            "id": r["id"],
            "codigo": r["codigo"],
            "ean": r["ean"],
            "dep": r["dep"],
            "departamento": r["departamento"],
            "descripcion": r["descripcion"],
            "Normal": r["normal"],
            "Oferta": r["oferta"],
            "cenefa": r["cenefa"],
            "desde": r["desde"],
            "hasta": r["hasta"],
            "sucursales": r["sucursales"],
            "tipo_cenefa": r["tipo_cenefa"],
            "fecha_carga": r["fecha_carga"],
            "lote_carga": r["lote_carga"],
            "usuario_carga": r["usuario_carga"],
        })

    return render_template(
        "historico_cenefas.html",
        datos=datos,
        filtro_codigo=filtro_codigo,
        filtro_tipo=filtro_tipo,
        filtro_lote=filtro_lote
    )

from sucursales.folder_mayorista_tucuman import (
    registrar_rutas_folder_mayorista_tucuman,
)

registrar_rutas_folder_mayorista_tucuman(
    compras_bp=compras_bp,
    procesar_archivo_cenefas=procesar_archivo_cenefas,
    guardar_cenefas_en_db=guardar_cenefas_en_db,
    get_db_connection=get_db_connection,
    limpiar_precio=limpiar_precio,
    formatear_precio_arg=formatear_precio_arg,
)

from sucursales.administrar_cenefas import (
    registrar_rutas_administrar_cenefas,
)


registrar_rutas_administrar_cenefas(
    compras_bp=compras_bp,
    get_db_connection=get_db_connection,
    limpiar_precio=limpiar_precio,
)
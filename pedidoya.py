import os
import json
import re
import pandas as pd
from flask import Blueprint, render_template, request, send_file
from openpyxl import load_workbook

pedidoya_bp = Blueprint("pedidoya", __name__)

ARCHIVO_TEMP = os.path.join(os.path.dirname(__file__), "pedidoya_procesado.xlsx")

ALIAS = {
    "sucursal": ["SUCURSALES", "Sucursal", "sucursales"],
    "descarga": ["descarga", "descargas"],
    "codigo": ["codigo", "cod", "Material", "material"],
    "descripcion": ["descripcion", "description", "desc"],
    "desde": ["desde", "from"],
    "hasta": ["hasta", "to"],
    "promo": ["PROMO","acciones", "acciones a realizar", "accion"]
}

def get_column_name(df, aliases, default):
    for col in df.columns:
        if str(col).strip().lower() in [a.lower() for a in aliases.get(default, [])]:
            return col
    return None

def parse_promo_for_df(promo_text):
    if not promo_text or str(promo_text).strip() == "":
        return "", ""
    promo_text = str(promo_text).strip()
    m = re.match(r"(\d+)x(\d+)", promo_text.lower())
    if m:
        return "mxn", 1
    number_match = re.search(r'\d+(\.\d+)?', promo_text)
    if number_match:
        num = float(number_match.group())
        number = int(num) if num.is_integer() else num
    else:
        number = ""
    if "%" in promo_text:
        action = "%"
    elif "$" in promo_text:
        action = "$"
    else:
        action = ""
    return action, number

def cantidad_from_promo(promo):
    if not promo or str(promo).strip() == "":
        return None

    promo = str(promo).lower().strip()
    m = re.match(r"(\d+)\s*x\s*(\d+)", promo)
    if m:
        return int(m.group(1))
    return None

def normalizar_columna(col):
    return (
        str(col)
        .strip()
        .lower()
        .replace(" ", "")
        .replace(".", "")
        .replace("-", "")
        .replace("_", "")
    )


def eliminar_columnas_pedidoya(df):
    columnas_a_eliminar = {
        "depto",
        "tipo",
        "idsku",
        "idproducto",
        "ean",
        "eanpeya",
    }

    columnas_drop = []

    for col in df.columns:
        if normalizar_columna(col) in columnas_a_eliminar:
            columnas_drop.append(col)

    print("Columnas eliminadas:", columnas_drop)

    return df.drop(columns=columnas_drop, errors="ignore")


#=========================================================
# Funcion para Promo llev
#=========================================================

def parse_promo_for_row(row):
    promo_text = row.get("PROMO", "")

    if not promo_text or str(promo_text).strip() == "":
        return "", ""

    promo_text = str(promo_text).strip()
    promo_lower = promo_text.lower()

    # Caso: "Llevando 2 Un", "Llevando 3 Un", etc.
    if re.search(r"llevando\s+\d+\s*un", promo_lower):
        normal = pd.to_numeric(row.get("Normal", None), errors="coerce")
        oferta = pd.to_numeric(row.get("Oferta", None), errors="coerce")

        if pd.notna(normal) and pd.notna(oferta):
            descuento = normal - oferta
        else:
            descuento = ""

        return "llev", descuento

    # Caso: 2x1, 3x2, etc.
    m = re.match(r"(\d+)\s*x\s*(\d+)", promo_lower)
    if m:
        return "mxn", 1

    number_match = re.search(r"\d+(\.\d+)?", promo_text)

    if number_match:
        num = float(number_match.group())
        number = int(num) if num.is_integer() else num
    else:
        number = ""

    if "%" in promo_text:
        action = "%"
    elif "$" in promo_text:
        action = "$"
    else:
        action = ""

    return action, number



def standardize_dataframe(df):
    standard_cols = ["Codigo", "Descripcion", "sucursal", "PROMO", "descarga",
                     "cantidad", "addcant", "rappel", "canper", "desde", "hasta"]

    new_data = {}

    for std_name in standard_cols:
        alias_key = std_name.lower()

        if std_name == "PROMO":
            orig_name = get_column_name(df, ALIAS, "promo")
        else:
            orig_name = get_column_name(df, ALIAS, alias_key)

        if orig_name and orig_name in df.columns:
            new_data[std_name] = df[orig_name]
        else:
            new_data[std_name] = [""] * len(df)

    df_std = pd.DataFrame(new_data)

    # Agregar Normal y Oferta SOLO para previsualización / cálculo
    for col in df.columns:
        col_norm = normalizar_columna(col)

        if col_norm == "normal":
            df_std["Normal"] = df[col]

        if col_norm == "oferta":
            df_std["Oferta"] = df[col]

    df_std[["accion", "descuento"]] = df_std.apply(
        lambda row: pd.Series(parse_promo_for_row(row)),
        axis=1
    )

    df_std["cantidad"] = pd.to_numeric(df_std["cantidad"], errors="coerce")
    cantidad_promo = df_std["PROMO"].apply(cantidad_from_promo)
    df_std["cantidad"] = df_std["cantidad"].fillna(cantidad_promo)
    df_std["cantidad"] = df_std["cantidad"].fillna(1).astype(int)

    df_std["addcant"] = pd.to_numeric(df_std["addcant"], errors="coerce").fillna(0).astype(int)
    df_std["rappel"] = pd.to_numeric(df_std["rappel"], errors="coerce").fillna(0).astype(int)
    df_std["canper"] = pd.to_numeric(df_std["canper"], errors="coerce").fillna(0).astype(int)
    df_std["sucursal"] = pd.to_numeric(df_std["sucursal"], errors="coerce").fillna(0).astype(int)

    return df_std


@pedidoya_bp.route("/", methods=["GET"])
def index():
    return render_template(
        "pedidoya.html",
        preview=False,
        columns=[],
        rows=[],
        table_data="[]"
    )


def combinar_celdas_por_descarga(archivo_excel):
    wb = load_workbook(archivo_excel)
    ws = wb.active

    columnas_a_combinar = ["descarga", "PROMO", "accion", "descuento"]

    headers = {
        cell.value: cell.column
        for cell in ws[1]
    }

    col_descarga = headers.get("descarga")

    if not col_descarga:
        wb.save(archivo_excel)
        return

    columnas_indices = [
        headers[col]
        for col in columnas_a_combinar
        if col in headers
    ]

    fila_inicio = 2
    valor_actual = ws.cell(row=fila_inicio, column=col_descarga).value

    for fila in range(3, ws.max_row + 2):
        valor = ws.cell(row=fila, column=col_descarga).value if fila <= ws.max_row else None

        if valor != valor_actual:
            fila_fin = fila - 1

            if fila_fin > fila_inicio:
                for col_idx in columnas_indices:
                    ws.merge_cells(
                        start_row=fila_inicio,
                        start_column=col_idx,
                        end_row=fila_fin,
                        end_column=col_idx
                    )

            fila_inicio = fila
            valor_actual = valor

    wb.save(archivo_excel)

def formatear_precio(valor):
    try:
        return "{:,.2f}".format(float(valor)).replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return valor

@pedidoya_bp.route("/preview", methods=["POST"])
def preview():
    archivo = request.files.get("file")

    if not archivo:
        return "No se cargó ningún archivo", 400

    df = pd.read_excel(archivo)
    df.columns = df.columns.astype(str).str.strip()

    df = eliminar_columnas_pedidoya(df)
    df = standardize_dataframe(df)

    # 👉 función para formatear
    def formatear_precio(valor):
        try:
            return "{:,.2f}".format(float(valor)).replace(",", "X").replace(".", ",").replace("X", ".")
        except:
            return valor

    # 👉 copia solo para mostrar
    df_view = df.copy()

    for col in ["Normal", "Oferta"]:
        if col in df_view.columns:
            df_view[col] = df_view[col].apply(formatear_precio)

    columns = list(df_view.columns)
    rows = df_view.fillna("").to_dict(orient="records")

    table_data = json.dumps(df.fillna("").to_dict(orient="records"), ensure_ascii=False)

    return render_template(
        "pedidoya.html",
        preview=True,
        columns=columns,
        rows=rows,
        table_data=table_data
    )

@pedidoya_bp.route("/descargar", methods=["POST"])
def descargar():
    table_data = request.form.get("table_data")

    if not table_data:
        return "No hay datos para descargar", 400

    data = json.loads(table_data)
    df = pd.DataFrame(data)

    df = df.drop(columns=["Normal", "Oferta"], errors="ignore")

    df.to_excel(ARCHIVO_TEMP, index=False)

    combinar_celdas_por_descarga(ARCHIVO_TEMP)

    return send_file(
        ARCHIVO_TEMP,
        as_attachment=True,
        download_name="pedido_ya.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )